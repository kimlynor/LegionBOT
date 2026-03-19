import io
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


JOB_COLORS = [
    'D9EAD3',  # 연초록
    'CFE2F3',  # 연파랑
    'FCE5CD',  # 연주황
    'EAD1DC',  # 연분홍
    'FFF2CC',  # 연노랑
    'D9D2E9',  # 연보라
    'D0E0E3',  # 연청록
    'F4CCCC',  # 연빨강
]

JOB_HEADER_COLORS = [
    'A8D08D',  # 초록
    '9DC3E6',  # 파랑
    'F4B183',  # 주황
    'C9A0C9',  # 분홍
    'FFD966',  # 노랑
    'B4A7D6',  # 보라
    '76A5AF',  # 청록
    'E06666',  # 빨강
]

TITLE_FILL  = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
TOTAL_FILL  = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')

CENTER = Alignment(horizontal='center', vertical='center')

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
MEDIUM_TOP = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='medium'), bottom=Side(style='thin'),
)

HEADERS = ['구분', '캐릭명', '아이템레벨', '전투력', '가능시간', '메모']
COL_WIDTHS = [8, 22, 16, 16, 25, 30]

DAY_MAP = {
    '월요일':           ['월'],
    '화요일':           ['화'],
    '수요일':           ['수'],
    '목요일':           ['목'],
    '금요일':           ['금'],
    '토요일':           ['토'],
    '일요일':           ['일'],
    '평일 전체 (월~금)': ['월', '화', '수', '목', '금'],
    '주말 (토~일)':      ['토', '일'],
    '수~일 무관':        ['수', '목', '금', '토', '일'],
    '월~화 무관':        ['월', '화'],
    '전체 무관':         ['월', '화', '수', '목', '금', '토', '일'],
}
DAY_ORDER = ['월', '화', '수', '목', '금', '토', '일']
DAY_FULL  = {'월': '월요일', '화': '화요일', '수': '수요일', '목': '목요일',
             '금': '금요일', '토': '토요일', '일': '일요일'}


def _set_cell(ws, row, col, value, font=None, fill=None, border=THIN, alignment=CENTER, height=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if border:    cell.border = border
    if alignment: cell.alignment = alignment
    if height:    ws.row_dimensions[row].height = height
    return cell


def _write_sheet(ws, applicants: list, party_purpose: str):
    num_cols = len(HEADERS)
    last_col_letter = chr(64 + num_cols)  # 'F'

    # 제목 행
    ws.merge_cells(f'A1:{last_col_letter}1')
    _set_cell(ws, 1, 1, f'파티 모집: {party_purpose}',
              font=Font(bold=True, size=13, color='FFFFFF'),
              fill=TITLE_FILL, height=28)

    if not applicants:
        ws.merge_cells(f'A2:{last_col_letter}2')
        _set_cell(ws, 2, 1, '지원자 없음', height=20)
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        return

    # 직업별 그룹핑
    groups = defaultdict(list)
    job_order = []
    for ap in applicants:
        job = ap['job']
        if job not in groups:
            job_order.append(job)
        groups[job].append(ap)

    job_color_map = {job: JOB_COLORS[i % len(JOB_COLORS)]              for i, job in enumerate(job_order)}
    job_hdr_map   = {job: JOB_HEADER_COLORS[i % len(JOB_HEADER_COLORS)] for i, job in enumerate(job_order)}

    row = 2
    total = len(applicants)

    for idx, job in enumerate(job_order):
        members = groups[job]

        # 직업 구분 헤더
        ws.merge_cells(f'A{row}:{last_col_letter}{row}')
        _set_cell(ws, row, 1, f'[ {job} ]  {len(members)}명',
                  font=Font(bold=True, size=11, color='FFFFFF'),
                  fill=PatternFill(start_color=job_hdr_map[job],
                                   end_color=job_hdr_map[job], fill_type='solid'),
                  border=MEDIUM_TOP, height=22)
        row += 1

        # 컬럼 헤더 (직업마다 반복)
        for col, h in enumerate(HEADERS, 1):
            _set_cell(ws, row, col, h,
                      font=Font(bold=True, color='FFFFFF', size=10),
                      fill=HEADER_FILL, height=20)
        row += 1

        # 데이터 행 (전투력 내림차순)
        data_fill = PatternFill(start_color=job_color_map[job],
                                end_color=job_color_map[job], fill_type='solid')
        for ap in members:
            values = [
                '부캐' if ap['is_sub'] else '메인',
                ap['char_name'],
                ap['combat_power'],
                ap['atool_score'],
                ap['available_time'] or '',
                ap['memo'] or '',
            ]
            for col, val in enumerate(values, 1):
                _set_cell(ws, row, col, val, fill=data_fill, height=20)
            row += 1

        # 직업 그룹 사이 빈 행 2줄 (마지막 그룹 제외)
        if idx < len(job_order) - 1:
            row += 2

    # 합계 행
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    _set_cell(ws, row, 1, f'총 {total}명',
              font=Font(bold=True, size=11),
              fill=TOTAL_FILL, height=22)

    # 열 너비
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def create_party_excel(applicants: list, party_purpose: str) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # 전체 시트
    ws_all = wb.active
    ws_all.title = '전체'
    _write_sheet(ws_all, applicants, party_purpose)

    # 요일별 시트
    for day_short in DAY_ORDER:
        day_applicants = [
            ap for ap in applicants
            if day_short in DAY_MAP.get(ap.get('available_days', ''), [])
        ]
        ws = wb.create_sheet(title=DAY_FULL[day_short])
        _write_sheet(ws, day_applicants, party_purpose)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
